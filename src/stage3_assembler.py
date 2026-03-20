import json
import os
import glob
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

COLUMN_MAP = {
    "nhs_number": "NHS Number",
    "patient_name": "Patient Name",
    "dob": "Date of Birth",
    "hospital_number": "Hospital Number",
    "consultant": "Consultant",
    "mdt_date": "MDT Date",
    "mrT_stage": "Baseline MRI: mrT Stage",
    "mrN_stage": "Baseline MRI: mrN Stage",
    "mrCRM_status": "Baseline MRI: CRM Status",
    "mrCRM_distance_mm": "Baseline MRI: CRM Distance (mm)",
    "mrEMVI_status": "Baseline MRI: EMVI Status",
    "mrMRF_status": "Baseline MRI: MRF Status",
    "tumour_height_cm": "Tumour Height from Anal Verge (cm)",
    "mri_date": "Baseline MRI Date",
    "is_restaging_mri": "Restaging MRI Present",
    "restaging_mrT_stage": "Restaging MRI: mrT Stage",
    "restaging_mrN_stage": "Restaging MRI: mrN Stage",
    "restaging_mrCRM_status": "Restaging MRI: CRM Status",
    "restaging_mrEMVI_status": "Restaging MRI: EMVI Status",
    "mrTRG": "Restaging MRI: Tumour Regression Grade (mrTRG)",
    "restaging_mri_date": "Restaging MRI Date",
    "ct_M_stage": "CT: M Stage",
    "ct_liver_mets": "CT: Liver Metastases",
    "ct_lung_mets": "CT: Lung Metastases",
    "ct_peritoneal_disease": "CT: Peritoneal Disease",
    "ct_date": "CT Date",
    "histology_type": "Histology Type",
    "histology_grade": "Histology Grade",
    "mmr_status": "MMR Status",
    "msi_status": "MSI Status",
    "kras_status": "KRAS Status",
    "nras_status": "NRAS Status",
    "braf_status": "BRAF Status",
    "her2_status": "HER2 Status",
    "pathology_date": "Pathology Date",
    "treatment_intent": "MDT Treatment Intent",
    "primary_treatment_modality": "Primary Treatment Modality",
    "planned_surgery": "Planned Surgery",
    "neoadjuvant_treatment": "Neoadjuvant Treatment Planned",
    "mdt_outcome_verbatim": "MDT Outcome (Verbatim)",
    "verification_required": "Human Verification Required",
    "verification_reasons": "Verification Reasons"
}

def assemble_excel(structured_files: list, excel_dir: str, json_dir: str) -> str:
    """
    Assemble Excel workbook from structured cases.
    
    Args:
        structured_files: List of paths to case_*_structured.json files
        excel_dir: Output directory for Excel file
        json_dir: Directory containing raw case JSON files (for token reversal)
    
    Returns:
        Path to generated Excel file
    """
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "MDT Data"
    ws2 = wb.create_sheet("Source Evidence")
    ws3 = wb.create_sheet("Flags & Notes")
    
    # Setup Sheet 1 Headers
    ordered_keys = list(COLUMN_MAP.keys())
    headers = [COLUMN_MAP[k] for k in ordered_keys]
    ws1.append(headers)
    
    # Banding Colors mapping by group (simplified heuristic for length)
    color_map = {
        "nhs_number": "F2F2F2", "mrT_stage": "DDEEFF", "restaging_mrT_stage": "DDFFEE",
        "ct_M_stage": "FFFADD", "histology_type": "EEFFDD", "treatment_intent": "FFE8CC",
        "verification_required": "FFE0E0"
    }
    current_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Note: Full color banding logic can be refined per column in a real setting, 
    # applying manually here per spec:
    for col_idx, col_key in enumerate(ordered_keys, 1):
        if col_key in ["nhs_number", "patient_name", "dob", "hospital_number", "consultant", "mdt_date"]:
            fill = PatternFill(start_color="F2F2F2", fill_type="solid")
        elif col_key in ["mrT_stage", "mrN_stage", "mrCRM_status", "mrCRM_distance_mm", "mrEMVI_status", "mrMRF_status", "tumour_height_cm", "mri_date", "is_restaging_mri"]:
            fill = PatternFill(start_color="DDEEFF", fill_type="solid")
        elif col_key in ["restaging_mrT_stage", "restaging_mrN_stage", "restaging_mrCRM_status", "restaging_mrEMVI_status", "mrTRG", "restaging_mri_date"]:
            fill = PatternFill(start_color="DDFFEE", fill_type="solid")
        elif col_key in ["ct_M_stage", "ct_liver_mets", "ct_lung_mets", "ct_peritoneal_disease", "ct_date"]:
            fill = PatternFill(start_color="FFFADD", fill_type="solid")
        elif col_key in ["histology_type", "histology_grade", "mmr_status", "msi_status", "kras_status", "nras_status", "braf_status", "her2_status", "pathology_date"]:
            fill = PatternFill(start_color="EEFFDD", fill_type="solid")
        elif col_key in ["treatment_intent", "primary_treatment_modality", "planned_surgery", "neoadjuvant_treatment", "mdt_outcome_verbatim"]:
            fill = PatternFill(start_color="FFE8CC", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFE0E0", fill_type="solid")
        ws1.cell(row=1, column=col_idx).fill = fill

    ws2.append(["Patient ID", "Patient Name", "MDT Date", "Field Name", "Clinical Label", "Extracted Value", "Source Text", "Confidence"])
    ws3.append(["Case Index", "Patient ID", "MDT Date", "Verification Reasons", "Notes"])
    
    # A1 comment 
    ws1["A1"].comment = openpyxl.comments.Comment("Amber rows require manual review against the Source Evidence sheet before this data is used clinically.", "System")
    ws1.freeze_panes = ws1['C2']

    # Load data and pre-sort by MDT date (DD/MM/YYYY)
    case_bundle = []

    def _mdt_key(date_str: str):
        try:
            day, month, year = date_str.split('/')
            return (int(year), int(month), int(day))
        except Exception:
            return (9999, 12, 31)

    for s_file in structured_files:
        with open(s_file, 'r', encoding='utf-8') as sf:
            s_data = json.load(sf)
        case_index = s_data.get('case_index', 0)
        
        # Try multiple naming conventions for raw file
        raw_file = None
        for candidate in [
            os.path.join(json_dir, f"case_{case_index:03d}.json"),
            os.path.join(json_dir, f"case_{case_index}_raw.json"),
        ]:
            if os.path.exists(candidate):
                raw_file = candidate
                break
        
        if not raw_file:
            print(f"⚠️ Warning: Could not find raw file for case {case_index:03d}")
            r_data = {"patient_identifiers": {}, "token_map": {}}
        else:
            with open(raw_file, 'r', encoding='utf-8') as rf:
                r_data = json.load(rf)
        
        case_bundle.append((s_data, r_data))

    for s_data, r_data in sorted(case_bundle, key=lambda x: _mdt_key(x[1].get('patient_identifiers', {}).get('mdt_date', ''))):
        case_index = s_data.get('case_index', 0)
        patient_identifiers = r_data.get('patient_identifiers', {})
        token_map = r_data.get('token_map', {})
        patient_id = patient_identifiers.get('nhs_number', '') or patient_identifiers.get('hospital_number', '')
        patient_name = patient_identifiers.get('patient_name', '')
        mdt_date = patient_identifiers.get('mdt_date', '')
        
        flattened = {}
        def _flatten(d):
            for k, v in d.items():
                if isinstance(v, dict) and 'value' in v:
                    raw_value = v.get('value')
                    source_txt = v.get('source_text')

                    # De-tokenise both values and source text so outputs are human-readable
                    if isinstance(raw_value, str):
                        for token, orig_val in token_map.items():
                            raw_value = raw_value.replace(token, str(orig_val))
                    if isinstance(source_txt, str):
                        for token, orig_val in token_map.items():
                            source_txt = source_txt.replace(token, str(orig_val))

                    flattened[k] = {"value": raw_value, "source_text": source_txt, "confidence": v.get('confidence')}
                elif isinstance(v, dict):
                    _flatten(v)
        _flatten(s_data)

        row = []
        for key in ordered_keys:
            if key in patient_identifiers:
                row.append(patient_identifiers.get(key))
            elif key == 'verification_required':
                row.append("YES" if s_data.get('verification_required') else "NO")
            elif key == 'verification_reasons':
                row.append(s_data.get('verification_reasons'))
            elif key == 'is_restaging_mri':
                row.append(s_data.get('is_restaging_mri', False))
            else:
                row.append(flattened.get(key, {}).get('value'))

        ws1.append(row)
        current_row_idx = ws1.max_row
        if s_data.get('verification_required'):
            amber_fill = PatternFill(start_color="FFC000", fill_type="solid")
            for col in range(1, len(row) + 1):
                ws1.cell(row=current_row_idx, column=col).fill = amber_fill

        for k, v in flattened.items():
            if v.get('value') is not None:
                new_row2 = [patient_id, patient_name, mdt_date, k, COLUMN_MAP.get(k, k), v.get('value'), v.get('source_text'), v.get('confidence')]
                ws2.append(new_row2)
                conf = v.get('confidence', '')
                c_cell = ws2.cell(row=ws2.max_row, column=8)
                if conf == 'high':
                    c_cell.fill = PatternFill(start_color="CCFFCC", fill_type="solid")
                elif conf == 'low':
                    c_cell.fill = PatternFill(start_color="FFC000", fill_type="solid")
                else:
                    c_cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")

        if s_data.get('verification_required'):
            ws3.append([case_index, patient_id, mdt_date, s_data.get('verification_reasons'), ""])

    # Auto-fit columns
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: 
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min((max_length + 2), 40)
            ws.column_dimensions[column].width = adjusted_width

    # Generate output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(excel_dir, f"MDT_Extraction_{timestamp}.xlsx")
    
    wb.save(output_file)
    return output_file
